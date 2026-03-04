# app/form_builder/form_export.py

import io
import csv
from datetime import datetime
from flask import send_file, make_response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class FormExporter:
    """
    Enterprise Data Export Engine
    Sözlük (Dict) listelerini alır ve şık formatlanmış dosyalara dönüştürür.
    """
    
    @staticmethod
    def _prepare_headers(columns):
        """Kolon tanımlarından başlıkları çıkarır"""
        return [col.get('label', col.get('name', '')) for col in columns]

    @staticmethod
    def _prepare_row(item, columns):
        """Her bir veriyi kolon sırasına göre hizalar ve temizler"""
        row = []
        for col in columns:
            val = item.get(col['name'], '')
            # Tarihleri veya objeleri string'e çevir
            if isinstance(val, datetime):
                val = val.strftime('%d.%m.%Y %H:%M')
            elif val is None:
                val = ''
            row.append(str(val))
        return row

    @classmethod
    def to_excel(cls, data_list, columns, filename="Rapor"):
        """Şık tasarımlı (Koyu başlıklar, otomatik sütun genişliği) Excel üretir"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Dışa Aktarım"

        # Başlık Stilleri
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4E73DF", end_color="4E73DF", fill_type="solid")
        alignment = Alignment(horizontal="left", vertical="center")

        headers = cls._prepare_headers(columns)
        ws.append(headers)

        # Başlık hücrelerini boya
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            # Sütun genişliğini kabaca ayarla
            ws.column_dimensions[cell.column_letter].width = 20

        # Verileri yaz
        for item in data_list:
            ws.append(cls._prepare_row(item, columns))

        # Hafızada tut ve Flask ile gönder
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        safe_filename = f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(out, download_name=safe_filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @classmethod
    def to_csv(cls, data_list, columns, filename="Rapor"):
        """Hızlı ve UTF-8 (Türkçe karakter destekli) CSV üretir"""
        si = io.StringIO()
        
        # UTF-8 BOM ekle ki Excel Türkçe karakterleri bozmasın!
        si.write('\ufeff') 
        
        writer = csv.writer(si, delimiter=';') # Türkiye standartı noktalı virgüldür
        
        writer.writerow(cls._prepare_headers(columns))
        for item in data_list:
            writer.writerow(cls._prepare_row(item, columns))

        output = make_response(si.getvalue())
        safe_filename = f"{filename}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        output.headers["Content-Disposition"] = f"attachment; filename={safe_filename}"
        output.headers["Content-type"] = "text/csv; charset=utf-8"
        return output